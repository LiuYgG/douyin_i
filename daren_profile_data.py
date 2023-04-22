"""
# FileName    : daren_profile_data.py
# CreateTime  : 2023 年 04 月 11 日
# Version     : 
# Description :
"""
import datetime

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import openpyxl
import time
from bases import daren_square_page
from bases import profiles_page
from bases import send_invitation_page
from bases import add_operate_window

class DarenProfiles():

    def __init__(self):
        # 创建一个 workbook 对象
        self.workbook = openpyxl.Workbook()
        # 创建一个 worksheet 对象
        self.worksheet = self.workbook.active
        # 指定服务器
        self.options = Options()
        # 服务器地址
        self.options.add_experimental_option("debuggerAddress", "127.0.0.1:5247")
        # 打开浏览器
        self.service = Service('drivers/112.0.5615.50_chromedriver.exe')
        self.d = webdriver.Chrome(service=self.service, options=self.options)
        self.data_list = []

    def run(self):
        self.setup_worksheet()
        self.get_daren_info()
        self.write_to_worksheet()
        self.save_workboot()

    def setup_worksheet(self):

        # 设置列名
        self.worksheet.cell(row=1, column=1, value='姓名')
        self.worksheet.cell(row=1, column=2, value='粉丝数')
        # self.worksheet.cell(row=1, column=3, value='联系方式')
        self.worksheet.cell(row=1, column=4, value='网址')

    def get_daren_info(self):

        nicke_name_elements = self.d.find_elements(By.CLASS_NAME, daren_square_page.daren_square_name_class)
        count = int(input("请输入你要的人数: "))
        start_index = int(input("请输入要开始的位置："))
        # 循环获取
        for n in range(start_index, len(nicke_name_elements)):
            if n == count:
                break
            try:
                self.daren = nicke_name_elements[n].click()
                # 获取所有打开的浏览器窗口
                handle_tab_all = self.d.window_handles
                # 获取当前浏览器的窗口
                self.d.current_window_handle
                self.d.switch_to.window(handle_tab_all[1])
                time.sleep(3)
                # 强制刷新
                self.d.execute_script("location.reload(true);")
                print("当前窗口已强制刷新")
                time.sleep(10)

                '''
                 联系方式相关操作
                '''
                try:
                    message_box = self.d.find_element(By.XPATH, profiles_page.profiles_message_box_xpath)
                    if message_box:
                        print("*" * 50)
                        print()
                        hide_btn = self.d.find_elements(By.CLASS_NAME, profiles_page.profiles_message_hide_class)
                        time.sleep(5)
                        try:
                            hide_btn[2].click()
                            hide_btn[3].click()
                            time.sleep(3)
                            print('-' * 30)
                            print(message_box.text)
                            print('-' * 30)
                        except Exception as h :
                            print(f"联系方式错误信息:{h}")
                except Exception as m:
                    if 'no such element: Unable to locate element: {"method":"xpath","selector":"//*[@id="app"]/div/div/div[1]/div[1]/div[2]/div[7]/div"}' in str(m):
                        time.sleep(1)
                        print(f"当前用户【 {self.d.find_element(By.XPATH, profiles_page.profiles_daren_name_xpath).text} 】没有设置联系方式信息")
                    else:
                        print(f"联系方式异常错误信息: {m}")
                time.sleep(10)

                '''
                若不支持邀约则抛出异常并继续下面的操作， 否则点击【发送邀约】
                '''
                try:
                    profiles_send = self.d.find_element(By.CLASS_NAME, profiles_page.profiles_send_class)
                    profiles_send_disabled = self.d.find_element(By.CLASS_NAME, profiles_page.profiles_send_disabled_class)
                    time.sleep(1)
                    if profiles_send_disabled :
                        time.sleep(1)
                        print("-" * 100)
                        print(f"当前用户【 {self.d.find_element(By.XPATH, profiles_page.profiles_daren_name_xpath).text} 】已禁止邀约")
                        # 关闭窗口
                        self.d.close()
                        self.d.switch_to.window((handle_tab_all[0]))
                        print("窗口已关闭")
                        time.sleep(5)  # 等待去下一个操作
                except Exception as s:
                    if 'no such element: Unable to locate element: {"method":"css selector","selector":".auxo-tooltip-disabled-compatible-wrapper"}' in str(s):
                        time.sleep(3)
                        print(f"当前用户【 {self.d.find_element(By.XPATH, profiles_page.profiles_daren_name_xpath).text} 】可以邀约")
                        time.sleep(3)
                        profiles_send.click()
                        print("已点击发送邀约按钮")
                        time.sleep(5)
                        self.d.find_element(By.CLASS_NAME, send_invitation_page.add_last_operate_class).click()
                        time.sleep(3)
                        print("已点击添加上次商品")
                        time.sleep(5)
                        # self.d.find_element(By.XPATH, send_invitation_page.send_invitation_btn_xpath).click()
                        self.d.find_element(By.XPATH, send_invitation_page.send_cancel_xpath).click()
                        print("已点击发送邀约")
                    else:
                        print(f'发送邀约异常错误信息：{s}')

                time.sleep(5)
                daren_name = self.d.find_element(By.XPATH, profiles_page.profiles_daren_name_xpath).text  # 达人的昵称
                daren_profile = self.d.find_element(By.CLASS_NAME, profiles_page.profiles_daren_fans_class).text # 达人的粉丝数
                # daren_contact = message_box.text  # 达人的联系方式

                print(f'姓名：{daren_name} | 粉丝数：{daren_profile}| '
                      # f'联系方式: {daren_contact} | '
                      f'网址：{self.d.current_url}')

                self.data_list.append({'姓名': daren_name,'粉丝数': daren_profile,
                                       # '联系方式': daren_contact,
                                       '网址': self.d.current_url})

                time.sleep(1)
                print(f"已联系: [ { n } ]个")
                print("-*" * 150)
                # 关闭窗口
                self.d.close()
                self.d.switch_to.window((handle_tab_all[0]))
                time.sleep(5)  # 等待去下一个操作
                # 向下滚动页面
                while True:
                    self.d.execute_script("window.scrollBy(0, 800);")
                    break
                time.sleep(5)
            except Exception as e:
                print(f'总循环错误信息: {e}')


    def write_to_worksheet(self):
        # 写入 Excel 表格
        for l, data in enumerate(self.data_list):
            self.worksheet.cell(row=l + 2, column=1, value=data['姓名'])
            self.worksheet.cell(row=l + 2, column=2, value=data['粉丝数'])
            # self.worksheet.cell(row=l + 2, column=3, value=data['联系方式'])
            self.worksheet.cell(row=l + 2, column=4, value=data['网址'])

    def save_workboot(self):
        # 保存数据
        self.workbook.save(f'datas/{time.strftime("%Y-%m-%d")}.xlsx')

if __name__ == '__main__':
    dd = DarenProfiles()
    dd.run()