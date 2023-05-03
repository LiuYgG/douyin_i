"""
# FileName    : daren_sen_and_contact.py
# CreateTime  : 2023 年 04 月 11 日
# Version     : 
# Description :
"""
import datetime
import os.path

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import openpyxl
import time
from bases import daren_square_page
from bases import profiles_page
from bases import send_invitation_page
from bases import add_operate_window

class DarenProfiles():

    def __init__(self):
        # 创建一个 workbook 对象
        self.workbook = openpyxl.Workbook()
        # 创建一个 worksheet 对象
        self.worksheet = self.workbook.active
        # 指定服务器
        self.options = Options()
        # 服务器地址
        self.options.add_experimental_option("debuggerAddress", "127.0.0.1:5247")
        # 打开浏览器
        self.service = Service('drivers/112.0.5615.50_chromedriver.exe')
        self.d = webdriver.Chrome(service=self.service, options=self.options)
        self.data_list = []

    def run(self):
        self.setup_worksheet()
        self.get_daren_info()
        self.write_to_worksheet()
        self.save_workboot()

    def setup_worksheet(self):

        # 设置列名
        self.worksheet.cell(row=1, column=1, value='姓名')
        self.worksheet.cell(row=1, column=2, value='粉丝数')
        self.worksheet.cell(row=1, column=3, value='联系方式')
        self.worksheet.cell(row=1, column=4, value='网址')

    def get_daren_info(self):

        nicke_name_elements = self.d.find_elements(By.CLASS_NAME, daren_square_page.daren_square_name_class)
        numbers = int(input("请输入你要的人数: "))
        start_index = int(input("请输入要开始的位置："))
        print("开始执行程序")
        # 循环获取
        for n in range(start_index, len(nicke_name_elements)):
            if n - start_index == numbers:
                print("程序执行完毕")
                break
            try:
                self.daren = nicke_name_elements[n].click()
                # 获取所有打开的浏览器窗口
                handle_tab_all = self.d.window_handles
                # 获取当前浏览器的窗口
                self.d.current_window_handle
                self.d.switch_to.window(handle_tab_all[1])
                time.sleep(3)
                # 强制刷新
                self.d.execute_script("location.reload(true);")
                print("*" * 150)
                print("当前窗口已强制刷新")
                time.sleep(10)
                """
                联系方式相关操作：
                    1.找到联系方式模块
                    2.判断联系方式是否存在，若不存在则输出 {用户没有设置联系方式}
                    3.找到联系方式隐藏按钮
                    4.遍历 message_box
                    5.判断 隐藏按钮
                """
                try:
                    message_box = self.d.find_element(By.XPATH, profiles_page.profiles_message_box_xpath)
                    if message_box:
                        hide_btn = self.d.find_elements(By.CLASS_NAME, profiles_page.profiles_message_hide_class)
                        time.sleep(5)
                        try:
                            hide_btn[2].click()
                            """
                            查看联系方式提示弹窗
                            """
                            time.sleep(5)
                            try:
                                message_toast_window = self.d.find_element(By.CLASS_NAME, profiles_page.profile_toast_window_class)
                                print("查看联系方式提示弹窗出现")
                                if message_toast_window < 0:
                                    print("无弹窗出现")
                            except Exception as w:
                                if message_toast_window:
                                    message_toast_window_sure = self.d.find_element(By.XPATH, profiles_page.profile_toast_window_sure_xpath)
                                    message_toast_window_sure.click()
                                    print("查看联系方式提示弹窗已点击查看")
                                elif 'no such element: Unable to locate element: {"method":"css selector","selector":".auxo-modal-body"}' in str(w):
                                    message_toast_window_sure.click()
                                elif 'stale element reference: element is not attached to the page document' in str(w):
                                    pass
                                elif "cannot access local variable 'message_toast_window_sure' where it is not associated with a value" in str(w):
                                    continue
                                print(f'查看方式提示弹窗异常信息：{w}')

                            time.sleep(5)
                            hide_btn[3].click()
                            time.sleep(3)
                            print(message_box.text)
                            print('-' * 30)
                        except Exception as h :
                            print(f"联系方式错误信息:{h}")
                except Exception as m:
                    if 'no such element: Unable to locate element: {"method":"xpath","selector":"//*[@id="app"]/div/div/div[1]/div[1]/div[2]/div[7]/div"}' in str(m):
                        time.sleep(1)
                        print(f"当前用户【 {self.d.find_element(By.XPATH, profiles_page.profiles_daren_name_xpath).text} 】没有设置联系方式信息")
                    else:
                        print(f"联系方式异常错误信息: {m}")
                time.sleep(10)

                '''
                发送邀约相关操作
                    1.找到发送邀约的两种状态
                    2.判断发送邀约的状态
                    3.通过 profiles_send_disable 判断是 待回复 or 禁止邀约，如是就提示 {当前用户不支持邀约}
                    4.如果支持邀约则提示 {支持邀约} 并点击按钮
                '''
                try:
                    profiles_send = self.d.find_element(By.CLASS_NAME, profiles_page.profiles_send_class)
                    profiles_send_disabled = self.d.find_element(By.CLASS_NAME, profiles_page.profiles_send_disabled_class)
                    time.sleep(1)
                    if profiles_send_disabled :
                        time.sleep(1)
                        print("-" * 100)
                        print(f"当前用户【 {self.d.find_element(By.XPATH, profiles_page.profiles_daren_name_xpath).text} 】不支持邀约")
                        # 关闭窗口
                        self.d.close()
                        self.d.switch_to.window((handle_tab_all[0]))
                        print("已关闭当前窗口")
                        continue
                except Exception as s:
                    if 'no such element: Unable to locate element: {"method":"css selector","selector":".auxo-tooltip-disabled-compatible-wrapper"}' in str(s):
                        time.sleep(3)
                        print(f"当前用户【 {self.d.find_element(By.XPATH, profiles_page.profiles_daren_name_xpath).text} 】可以邀约")
                        time.sleep(3)
                        profiles_send.click()
                        print("已点击发送邀约按钮")
                        time.sleep(5)
                        self.d.find_element(By.CLASS_NAME, send_invitation_page.add_last_operate_class).click()
                        time.sleep(3)
                        print("已点击添加上次商品")
                        time.sleep(5)
                        self.d.find_element(By.XPATH, send_invitation_page.send_invitation_btn_xpath).click()
                        print("已点击发送邀约")
                    else:
                        print(f'发送邀约异常错误信息：{s}')

                time.sleep(5)
                daren_name = self.d.find_element(By.XPATH, profiles_page.profiles_daren_name_xpath).text  # 达人的昵称
                daren_profile = self.d.find_element(By.CLASS_NAME, profiles_page.profiles_daren_fans_class).text # 达人的粉丝数
                daren_contact = message_box.text # 达人联系方式


                print(f'姓名：{daren_name} | 粉丝数：{daren_profile}| '
                      f'联系方式: {daren_contact} | '
                      f'网址：{self.d.current_url}')

                self.data_list.append({'姓名': daren_name,'粉丝数': daren_profile,
                                       '联系方式': daren_contact,
                                       '网址': self.d.current_url})

                time.sleep(1)
                print(f"已联系: [ { n - start_index + 1 } ]个")
                print("*" * 150)
                # 关闭窗口
                self.d.close()
                self.d.switch_to.window((handle_tab_all[0]))
                # 等待5秒钟去下一个操作
                time.sleep(5)
                # 向下滚动页面
                while True:
                    self.d.execute_script("window.scrollBy(0, 800);")
                    break
                time.sleep(5)
            except Exception as e:
                if 'no such element: Unable to locate element: {"method":"xpath","selector":"//*[@id="app"]/div/div/div[1]/div[1]/div[2]/div[2]/div"}' in str(e):
                    pass
                elif 'element click intercepted: Element' in str(e):
                    pass
                elif "local variable 'message_box' referenced before assignment" in str(e):
                    pass
                elif 'stale element reference: element is not attached to the page document' in str(e):
                    pass
                else:
                    print(f'总循环错误信息: {e}')


    def write_to_worksheet(self):
        # 写入 Excel 表格
        for l, data in enumerate(self.data_list):
            self.worksheet.cell(row=l + 2, column=1, value=data['姓名'])
            self.worksheet.cell(row=l + 2, column=2, value=data['粉丝数'])
            self.worksheet.cell(row=l + 2, column=3, value=data['联系方式'])
            self.worksheet.cell(row=l + 2, column=4, value=data['网址'])

    def save_workboot(self):
        # 保存数据
        save_path = f'datas/{time.strftime("%Y-%m-%d")}.xlsx'
        self.workbook.save(save_path)
        # print('数据保存位置: ' + os.path.dirname(save_path))
if __name__ == '__main__':
    dd = DarenProfiles()
    dd.run()